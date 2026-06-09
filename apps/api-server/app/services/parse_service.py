"""Document parsing service — lightweight format-specific parsers.

Uses per-format libraries (pymupdf, python-docx, openpyxl, xlrd) instead of the
heavy unstructured library which pulls in PyTorch, transformers, ONNX, etc.

Handles: PDF, DOCX, XLSX, XLS, TXT, CSV, MD, and common image formats.
Legacy .doc (pre-2007 Word) is NOT supported — users should convert to .docx.
"""

from io import BytesIO
from pathlib import Path

from app.core.config import settings

# File extensions we can parse
SUPPORTED_EXTENSIONS: set[str] = {
    ".pdf", ".docx", ".xlsx", ".xls",
    ".txt", ".csv", ".md",
    ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp",
}

# Extensions that produce meaningful text
TEXT_PRODUCING_EXTENSIONS: set[str] = {
    ".pdf", ".docx", ".xlsx", ".xls", ".txt", ".csv", ".md",
}


def estimate_tokens(text: str) -> int:
    """Rough token count estimate. ~4 chars per token for English/Chinese mixed."""
    return max(1, len(text) // 4)


# ── per-format parsers (all work on bytes, no temp files needed) ──────────


def _parse_pdf(content: bytes) -> str:
    """Extract text from PDF using pymupdf (fitz) — C-backed, no ML deps."""
    import fitz  # pymupdf
    doc = fitz.open(stream=content, filetype="pdf")
    try:
        return "\n".join(page.get_text() for page in doc)
    finally:
        doc.close()


def _parse_docx(content: bytes) -> str:
    """Extract text from DOCX using python-docx — pure-XML parser."""
    from docx import Document
    doc = Document(BytesIO(content))
    return "\n".join(p.text for p in doc.paragraphs if p.text)


def _parse_xlsx(content: bytes) -> str:
    """Extract text from XLSX using openpyxl in read-only mode (low memory)."""
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    parts = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        parts.append(f"[Sheet: {sheet_name}]")
        for row in ws.iter_rows():
            parts.append("\t".join(str(cell.value or "") for cell in row))
    wb.close()
    return "\n".join(parts)


def _parse_xls(content: bytes) -> str:
    """Extract text from legacy XLS using xlrd — pure-Python, no deps."""
    import xlrd
    wb = xlrd.open_workbook(file_contents=content)
    parts = []
    for sheet in wb.sheets():
        parts.append(f"[Sheet: {sheet.name}]")
        for row_idx in range(sheet.nrows):
            parts.append("\t".join(
                str(sheet.cell_value(row_idx, col_idx))
                for col_idx in range(sheet.ncols)
            ))
    return "\n".join(parts)


def _parse_text(content: bytes) -> str:
    """Decode plain text — tries UTF-8, then GBK, then latin-1 fallback."""
    for encoding in ("utf-8", "gbk", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


# ── public API ─────────────────────────────────────────────────────────────


def parse_document(file_content: bytes, file_name: str) -> str:
    """Parse a document and return extracted text.

    Uses lightweight per-format libraries — no PyTorch / CUDA / ONNX.

    Raises:
        ValueError: if file extension is not in SUPPORTED_EXTENSIONS.
        RuntimeError: if parsing fails.
    """
    ext = Path(file_name).suffix.lower()

    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported file type: {ext}")

    # Images — not yet parseable in V1
    if ext in (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"):
        return f"[Image file: {file_name} — OCR not yet supported]"

    try:
        if ext == ".pdf":
            return _parse_pdf(file_content)
        elif ext == ".docx":
            return _parse_docx(file_content)
        elif ext == ".xlsx":
            return _parse_xlsx(file_content)
        elif ext == ".xls":
            return _parse_xls(file_content)
        elif ext in (".txt", ".csv", ".md"):
            return _parse_text(file_content)
        else:
            return f"[Unsupported file type: {ext}]"
    except Exception as e:
        raise RuntimeError(f"Failed to parse {file_name}: {e}") from e


def build_injection_text(parsed_text: str, file_name: str) -> str:
    """Build the text block to inject into the LLM message.

    For small documents: inject full text.
    For large documents: inject truncated text with a note.
    """
    token_estimate = estimate_tokens(parsed_text)

    if token_estimate <= settings.MAX_FILE_SIZE_MB * 1024 * 256:
        # Full injection
        return (
            f"\n\n--- Content of {file_name} ---\n"
            f"{parsed_text}\n"
            f"--- End of {file_name} ---"
        )
    else:
        # Truncated injection
        max_chars = settings.MAX_FILE_SIZE_MB * 1024 * 256
        truncated = parsed_text[:max_chars]
        return (
            f"\n\n--- Content of {file_name} (truncated) ---\n"
            f"{truncated}\n"
            f"... [Document truncated. Full document is approximately "
            f"{token_estimate:,} tokens.]\n"
            f"--- End of {file_name} ---"
        )
